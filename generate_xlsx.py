import argparse
import math
import re
import typing
import os

import numpy
import pandas

from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Border, Fill, Font, PatternFill, Side, alignment, borders, colors, fills, numbers
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet


# region Configuration

F_PATH = os.path.dirname(__file__)
F_NAME = f"{F_PATH}\\Pi.xlsx"
F_BASE = os.path.basename(F_NAME)

HIGH_NUMBER_MULTIPLIER = 3
OUTDATED = {  # outdated or only available since
    # "UP_CANN": "4.30",  # example
    # "UP_CANNX": "5.50",
    "UP_RBSUIT": "4.40",
    "UP_MFIRE": "5.00",
    "PROC_EXH": "5.50",
    "UP_CANNX": "5.50",
    "UP_UNW": "5.50",
    "UP_EXSUB": "5.50",
}
RE_LANGUAGE = re.compile("\(([A-Za-z1-9-]+)\)")
URL = "https://github.com/zencq/Pi"
VERSION = "4.30"

# endregion

# region Data

PRODUCT = [
    "PROC_LOOT",
    "PROC_HIST",
    "PROC_BIO",
    "PROC_FOSS",
    "PROC_PLNT",
    "PROC_TOOL",
    "PROC_FARM",
    "PROC_SEA",
    "PROC_FEAR",
    "PROC_SALV",
    "PROC_BONE",
    "PROC_DARK",
    "PROC_STAR",
    "PROC_EXH",
]

# order here defines how it will be in the generated file
TECHNOLOGY = {
    "Suit": {
        "UP_ENGY": ["X", "3"],
        "UP_HAZ": ["X"],
        "UP_JET": ["X", "4"],
        "UP_SHLD": ["X", "4"],
        "UP_SNSUIT": [""],
        "UP_RBSUIT": [""],
        "UP_UNW": ["3"],
        "UP_RAD": ["3"],
        "UP_TOX": ["3"],
        "UP_COLD": ["3"],
        "UP_HOT": ["3"],
    },

    "Ship": {
        "UP_PULSE": ["X", "4"],
        "UP_LAUN": ["X", "4"],
        "UP_HYP": ["X", "4"],
        "UP_S_SHL": ["X", "4"],
        "UP_SGUN": ["X", "4"],
        "UP_SLASR": ["X", "4"],
        "UP_SSHOT": ["X", "4"],
        "UP_SMINI": ["X", "4"],
        "UP_SBLOB": ["X", "4"],
    },
    "AlienShip": {
        "UA_PULSE": ["4"],
        "UA_LAUN": ["4"],
        "UA_HYP": ["4"],
        "UA_S_SHL": ["4"],
        "UA_SGUN": ["4"],
        "UA_SLASR": ["4"],
    },

    "Weapon": {
        "UP_LASER": ["X", "4"],
        "UP_SCAN": ["X", "4"],
        "UP_BOLT": ["X", "4"],
        "UP_GREN": ["X", "4", "1"],
        "UP_TGREN": ["X", "4"],
        "UP_RAIL": ["X", "4"],
        "UP_SHOT": ["X", "4"],
        "UP_SMG": ["X", "4"],
        "UP_CANN": ["X", "4"],
        "UP_SENGUN": [""],
    },

    "Exocraft": {
        "UP_EXGUN": ["4"],
        "UP_EXLAS": ["4"],
        "UP_BOOST": ["4"],
        "UP_EXENG": ["4"],
    },
    "Submarine": {
        "UP_EXSUB": ["4"],
        "UP_SUGUN": ["4"],
    },
    "Mech": {
        "UP_MCLAS": ["4"],
        "UP_MCGUN": ["4"],
        "UP_MCENG": ["4"],
        "UP_MFIRE": ["4"],
    },

    "Freighter": {
        "UP_FRHYP": ["4"],
        "UP_FRSPE": ["4"],
        "UP_FRFUE": ["4"],
        "UP_FRCOM": ["4"],
        "UP_FRTRA": ["4"],
        "UP_FREXP": ["4"],
        "UP_FRMIN": ["4"],
    },
}
TECHNOLOGY_INDEX = [item_id for _, items in TECHNOLOGY.items() for item_id in items]

# endregion

# region Styles

ALIGNMENT_CENTER = Alignment(horizontal=alignment.horizontal_alignments[2])  # center
ALIGNMENT_LEFT = Alignment(horizontal=alignment.horizontal_alignments[1])  # left

BORDER_BOTTOM = Border(bottom=Side(border_style=borders.BORDER_MEDIUM, color=colors.BLACK))
BORDER_MIXED = Border(right=Side(border_style=borders.BORDER_DASHED, color=colors.BLACK), bottom=Side(border_style=borders.BORDER_MEDIUM, color=colors.BLACK))
BORDER_RIGHT = Border(right=Side(border_style=borders.BORDER_DASHED, color=colors.BLACK))

COLOR_COMMODITY = ("ffe8b3", "ffc94d", "ffbc25")  # gold
COLOR_FLORA = ("c8fab7", "7ef457", "8af567")  # green
COLOR_METAL = ("deddd3", "b3af98", "dbd9ce")  # gray
COLOR_SPECIAL = ("cdb3ff", "8b4dff", "be9bff")  # purple

FILL_RED = PatternFill(fill_type=fills.FILL_SOLID, start_color=colors.COLOR_INDEX[2])  # red

FONT_BOLD = Font(bold=True)

# endregion

# region Translation

LANGUAGES = [  # order defined by game
    "Name (en)",
    "Name (fr)",
    "Name (it)",
    "Name (de)",
    "Name (es)",
    "Name (ru)",
    "Name (pl)",
    "Name (nl)",
    "Name (pt)",
    "Name (es-419)",
    "Name (pt-BR)",
    "Name (ja)",
    "Name (zh-Hans)",
    "Name (zh-Hant)",
    "Name (ko)",
]

TRANSLATION = {
    # Inventories

    "AlienShip": "Starship (Living)",
    "Ship": "Starship (Regular)",
    "Suit": "Exosuit",
    "Weapon": "Multi-Tool",
    "Exocraft": "Exocraft",
    "Submarine": "Submarine (Nautilon)",
    "Mech": "Mech (Minotaur)",
    "Freighter": "Freighter",

    # Products

    "PROC_LOOT": "Unearthed Treasure",
    "PROC_HIST": "Historical Document",
    "PROC_BIO": "Biological Sample",
    "PROC_FOSS": "Fossil Sample",
    "PROC_PLNT": "Delicate Flora",
    "PROC_TOOL": "Lost Artifact",
    "PROC_FARM": "Delicate Flora",
    "PROC_SEA": "Aquatic Treasure",
    "PROC_FEAR": "Terrifying Sample",
    "PROC_SALV": "Salvaged Scrap",
    "PROC_BONE": "Excavated Bones",
    "PROC_DARK": "Terrifying Sample",
    "PROC_STAR": "Ancient Skeleton",
    "PROC_EXH": "Exhibit Fossil Sample",

    # Stats

    "Weapon_Laser_Damage": "Damage",
    "Weapon_Laser_Mining_Speed": "Mining Speed",
    "Weapon_Laser_HeatTime": "Heat Dispersion",
    "Weapon_Laser_ReloadTime": "Overheat Downtime",
    "Weapon_Laser_Drain": "Fuel Efficiency",
    "Weapon_Laser_ChargeTime": "Time to Full Power",
    "Weapon_Projectile_Damage": "Damage",
    "Weapon_Projectile_Rate": "Fire Rate",
    "Weapon_Projectile_ClipSize": "Clip Size",
    "Weapon_Projectile_ReloadTime": "Reload Time",
    "Weapon_Projectile_MaximumCharge": "Ion Spheres Created",
    "Weapon_Projectile_BurstCap": "Shots Per Burst",
    "Weapon_Projectile_BurstCooldown": "Burst Cooldown",
    "Weapon_ChargedProjectile_ChargeTime": "Charging Speed",
    "Weapon_ChargedProjectile_ExtraSpeed": "Ion Sphere Speed",
    "Weapon_Grenade_Damage": "Damage",
    "Weapon_Grenade_Radius": "Explosion Radius",
    "Weapon_Grenade_Speed": "Projectile Velocity",
    "Weapon_Grenade_Bounce": "Bounce Potential",
    "Weapon_Scan_Radius": "Scan Radius",
    "Weapon_Scan_Discovery_Creature": "Fauna Analysis Rewards",
    "Weapon_Scan_Discovery_Flora": "Flora Analysis Rewards",
    "Weapon_Scan_Discovery_Mineral": "Mineral Analysis Rewards",

    "Suit_Armour_Health": "Core Health",
    "Suit_Armour_Shield_Strength": "Shield Strength",
    "Suit_Energy": "Life Support Tanks",
    "Suit_Energy_Regen": "Solar Panel Power",
    "Suit_Protection_Cold": "Cold Protection",
    "Suit_Protection_Heat": "Heat Protection",
    "Suit_Protection_Toxic": "Toxic Protection",
    "Suit_Protection_Radiation": "Radiation Protection",
    "Suit_Underwater": "Oxygen Tank",
    "Suit_DamageReduce_Cold": "Cold Damage Shielding",
    "Suit_DamageReduce_Heat": "Heat Damage Shielding",
    "Suit_DamageReduce_Radiation": "Radiation Damage Shielding",
    "Suit_DamageReduce_Toxic": "Toxic Damage Shielding",
    "Suit_Protection_HeatDrain": "Heat Resistance",
    "Suit_Protection_ColdDrain": "Cold Resistance",
    "Suit_Protection_ToxDrain": "Toxic Resistance",
    "Suit_Protection_RadDrain": "Radiation Resistance",
    "Suit_Stamina_Strength": "Sprint Distance",
    "Suit_Stamina_Recovery": "Sprint Recovery Time",
    "Suit_Jetpack_Tank": "Jetpack Tanks",
    "Suit_Jetpack_Drain": "Fuel Efficiency",
    "Suit_Jetpack_Refill": "Recharge Rate",
    "Suit_Jetpack_Ignition": "Initial Boost Power",

    "Ship_Weapons_Guns_Damage": "Damage",
    "Ship_Weapons_Guns_Rate": "Fire Rate",
    "Ship_Weapons_Guns_HeatTime": "Heat Dispersion",
    "Ship_Weapons_Lasers_Damage": "Damage",
    "Ship_Weapons_Lasers_HeatTime": "Heat Dispersion",
    "Ship_Weapons_ShieldLeech": "Shield recharge on impact",
    "Ship_Armour_Shield_Strength": "Shield Strength",
    "Ship_Hyperdrive_JumpDistance": "Hyperdrive Range",
    "Ship_Hyperdrive_JumpsPerCell": "Warp Cell Efficiency",
    "Ship_Launcher_TakeOffCost": "Launch Cost",
    "Ship_Launcher_AutoCharge": "Automatic Recharging",
    "Ship_PulseDrive_MiniJumpFuelSpending": "Pulse Drive Fuel Efficiency",
    "Ship_Boost": "Boost",
    "Ship_Maneuverability": "Maneuverability",
    "Ship_BoostManeuverability": "Boost Maneuverability",

    "Freighter_Hyperdrive_JumpDistance": "Hyperdrive Range",
    "Freighter_Hyperdrive_JumpsPerCell": "Warp Cell Efficiency",
    "Freighter_Fleet_Speed": "Expedition Speed",
    "Freighter_Fleet_Fuel": "Expedition Efficiency",
    "Freighter_Fleet_Combat": "Expedition Defenses",
    "Freighter_Fleet_Trade": "Expedition Trade Ability",
    "Freighter_Fleet_Explore": "Expedition Scientific Ability",
    "Freighter_Fleet_Mine": "Expedition Mining Ability",

    "Vehicle_EngineFuelUse": "Fuel Usage",
    "Vehicle_EngineTopSpeed": "Top Speed",
    "Vehicle_BoostSpeed": "Boost Power",
    "Vehicle_BoostTanks": "Boost Tank Size",
    "Vehicle_SubBoostSpeed": "Acceleration",
    "Vehicle_LaserDamage": "Mining Laser Power",
    "Vehicle_LaserHeatTime": "Mining Laser Efficiency",
    "Vehicle_GunDamage": "Damage",
    "Vehicle_GunHeatTime": "Weapon Power Efficiency",
    "Vehicle_GunRate": "Rate of Fire",

    # Technologies

    "UP_COLD": "Cold Protection",
    "UP_ENGY": "Life Support",
    "UP_HAZ": "Hazard Protection",
    "UP_HOT": "Heat Protection",
    "UP_JET": "Movement System",
    "UP_RAD": "Radiation Protection",
    "UP_SHLD": "Defence Systems",
    "UP_SNSUIT": "Sentinel Exosuit Fragment",
    "UP_RBSUIT": "Rebuilt Exosuit Module",
    "UP_TOX": "Toxic Protection",
    "UP_UNW": "Aeration Membrane",

    "UP_HYP": "Hyperdrive",
    "UP_LAUN": "Launch Thruster",
    "UP_PULSE": "Pulse Engine",
    "UP_S_SHL": "Deflector Shield",
    "UP_SBLOB": "Cyclotron Ballista",
    "UP_SGUN": "Photon Cannon",
    "UP_SLASR": "Phase Beam",
    "UP_SMINI": "Infra-Knife Accelerator",
    "UP_SSHOT": "Positron Ejector",

    "UA_HYP": "Singularity Cortex",
    "UA_LAUN": "Neural Assembly",
    "UA_PULSE": "Pulsing Heart",
    "UA_S_SHL": "Scream Suppressor",
    "UA_SGUN": "Spewing Vents",
    "UA_SLASR": "Grafted Eyes",

    "UP_BOLT": "Boltcaster",
    "UP_CANN": "Neutron Cannon",
    "UP_GREN": "Plasma Launcher",
    "UP_LASER": "Mining Beam",
    "UP_RAIL": "Blaze Javelin",
    "UP_SCAN": "Analysis Visor",
    "UP_SENGUN": "Sentinel Weapons Shard",
    "UP_SHOT": "Scatter Blaster",
    "UP_SMG": "Pulse Spitter",
    "UP_TGREN": "Geology Cannon",

    "UP_BOOST": "Exocraft Boosters",
    "UP_EXENG": "Fusion Engine",
    "UP_EXGUN": "Mounted Cannon",
    "UP_EXLAS": "Exocraft Mining Laser",

    "UP_EXSUB": "Humboldt Drive",
    "UP_SUGUN": "Nautilon Cannon",

    "UP_MCENG": "Daedalus Engine",
    "UP_MCGUN": "Minotaur Cannon",
    "UP_MCLAS": "Minotaur Laser",
    "UP_MFIRE": "Minotaur Flamethrower",

    "UP_FRCOM": "Expedition Defenses Control",
    "UP_FREXP": "Expedition Scientific Control",
    "UP_FRFUE": "Expedition Efficiency Control",
    "UP_FRHYP": "Freighter Hyperdrive",
    "UP_FRMIN": "Expedition Mining Control",
    "UP_FRSPE": "Expedition Speed Control",
    "UP_FRTRA": "Expedition Trade Control",
}


def _(original: str) -> str:
    return TRANSLATION.get(original, original)


# endregion


class Pi():
    # region Property

    def __init__(self):
        super().__init__()

        # argparse
        self.debug = False
        self._language = LANGUAGES[1]  # english. name of column to use for the procedural names

        # properties to work with
        self.link_settings = {
            "column_product": -1,
            "column_technology": -1,
            "row_first": -1,
        }
        self.workbook = Workbook()

    @property
    def language(self):
        return self._language

    @language.setter
    def language(self, value):
        self._language = [l for l in LANGUAGES if value in l][0]

    @language.deleter
    def language(self):
        del self._language

    # endregion

    def add_products(self) -> None:
        product_sheet = self.sheet_create("Product")

        # insert header
        product_sheet.append(["", "Seed", "Perfection", "Age", "Value"])

        # insert data
        for item_name in PRODUCT:
            print("Product", item_name)  # to show progress
            product = self.get_product_with_pandas(item_name)
            self.insert_product(product_sheet, item_name, product)
            self.insert_in_overview_sheet("Product", item_name)

        self.sheet_autofit_column_width(product_sheet)
        self.column_3_set_border_right(product_sheet)

    def add_technologies(self) -> None:
        for inventory_type, items in TECHNOLOGY.items():
            inventory_translation = _(inventory_type)
            inventory_sheet = self.sheet_create(inventory_translation)

            for item_id, qualities in items.items():
                print(inventory_type, item_id)  # to show progress
                technologies = self.get_technology_with_pandas(inventory_type, item_id, qualities)
                anchor_coordinate = self.insert_technology(inventory_sheet, item_id, technologies)
                self.insert_in_overview_sheet(inventory_translation, item_id, target_coordinate=anchor_coordinate)

            self.sheet_autofit_column_width(inventory_sheet)
            self.column_3_set_border_right(inventory_sheet)

    @staticmethod
    def get_quality_information(item_name: str) -> tuple:
        # C -
        # S Supreme
        # X Illegal
        # ? Forbidden (Sentinel)
        # ! Salvaged (Autophage)

        quality = item_name[-1]

        # more general but defined case first
        if quality == "X":
            return "Illegal Upgrade (X)", COLOR_METAL

        # only used for UP_GREN and there C class is 1
        if quality == "1":
            return "Upgrade (1)", COLOR_FLORA

        # special cases in order of appearance
        if item_name in ["UP_SNSUIT", "UP_SENGUN"]:
            return "Forbidden Upgrade (?)", COLOR_SPECIAL

        if item_name == "UP_RBSUIT":
            return "Salvaged Upgrade (!)", COLOR_SPECIAL

        # no special case, therefore it can only be this
        return f"Supreme Upgrade ({quality})", COLOR_COMMODITY

    def write_xlsx(self) -> None:
        self.workbook.save(F_NAME)

    # region Helper for OpenPyXL

    @staticmethod
    def cell_set_fill_color(cell: Cell, fill: Fill) -> None:
        if not isinstance(fill, Fill):
            fill = PatternFill(fill_type=fills.FILL_SOLID, start_color=fill)

        cell.fill = fill

    @staticmethod
    def cell_set_hyperlink(cell: Cell, hyperlink: str) -> None:
        cell.hyperlink = hyperlink
        cell.style = "Hyperlink"

    @staticmethod
    def column_3_set_border_right(sheet: Worksheet) -> None:
        for row in sheet.iter_rows(min_row=sheet.min_row, max_row=sheet.max_row - 1, min_col=3, max_col=3):
            for cell in row:
                cell.border = BORDER_MIXED if cell.border.bottom.style else BORDER_RIGHT

    @staticmethod
    def row_max_set_border_bottom(sheet: Worksheet) -> None:
        for column in sheet.iter_cols(min_col=sheet.min_row, max_col=9, min_row=sheet.max_row, max_row=sheet.max_row):  # max_col not guaranteed 9
            for cell in column:
                cell.border = BORDER_MIXED if cell.border.right.style else BORDER_BOTTOM

    @staticmethod
    def row_max_set_fill_color(sheet: Worksheet, fill: Fill) -> None:
        if not isinstance(fill, Fill):
            fill = PatternFill(fill_type=fills.FILL_SOLID, start_color=fill)

        for column in sheet.iter_cols(min_col=sheet.min_column, max_col=9, min_row=sheet.max_row, max_row=sheet.max_row):  # max_col not guaranteed 9
            for cell in column:
                cell.fill = fill

    @staticmethod
    def row_max_merge_columns(sheet: Worksheet, start_column: int, end_column: int) -> None:
        sheet.merge_cells(start_row=sheet.max_row, start_column=start_column, end_row=sheet.max_row, end_column=end_column)

    @staticmethod
    def sheet_autofit_column_width(sheet: Worksheet) -> None:
        for column in sheet.columns:
            letter = column[0].column_letter
            sheet.column_dimensions[letter].width = max(len(str(cell.value)) for cell in column if "in game version" not in str(cell.value)) + 1

    def sheet_create(self, name: str) -> Worksheet:
        index = 0 if self.debug else None  # insert at first position to see it w/o clicking
        return self.workbook.create_sheet(name, index=index)

    # endregion

    # region Helper for Pandas

    def get_best_per_stat(self, dataframe_source: pandas.DataFrame) -> typing.Tuple[pandas.DataFrame, int, bool]:
        columns = self.get_stat_column_names_from_dataframe(dataframe_source)
        columns_max = [dataframe_source[column].max() for column in columns]

        if all(dataframe_source[column].min() == columns_max[i] for i, column in enumerate(columns)):
            return dataframe_source.tail(1), len(columns), True

        result = pandas.DataFrame()

        # add the absolute best per stat. taken from data to include all stats
        for column in columns:
            best = dataframe_source.sort_values(by=[column, "Perfection", "Seed"], ascending=False).head(1)
            result = pandas.concat([result, best])

        # drop duplicates
        result.drop_duplicates(subset="Seed", inplace=True)

        # only keep those with the highest number of perfect stats when there are any with more than one
        subset = {}
        max_stats_per_seed = 0

        for j, column in enumerate(columns):
            for k in range(len(result)):
                row = result.iloc[k]
                is_perfect = [row[c] == columns_max[i] for i, c in enumerate(columns)]
                is_perfect_sum = sum(is_perfect)

                max_stats_per_seed = max(max_stats_per_seed, sum(not math.isnan(row[c]) for c in columns))

                if is_perfect[j] and is_perfect_sum > subset.get(column, (column, 0))[1]:  # get a default tuple if necessary
                    subset[column] = (row["Seed"], is_perfect_sum)

        return result[result["Seed"].isin(value[0] for value in subset.values())], max_stats_per_seed, False

    @staticmethod
    def get_stat_column_names_from_dataframe(dataframes: pandas.DataFrame) -> typing.Set[str]:
        if isinstance(dataframes, pandas.DataFrame):
            dataframes = [dataframes]

        result = set()
        for dataframe in dataframes:
            result.update({column for column in dataframe if column not in ["Seed", "Perfection"] and not column.startswith("Name")})

        return sorted(result)

    # build meta data similar to the NMS.py mod script to be able to recalaculate perfection for Suit_Jetpack_Ignition Hi/Lo desirability
    def prepare_ignition_meta(self, dataframe: pandas.DataFrame) -> typing.Dict[str, typing.List[float]]:
        meta =  {
            stat: [dataframe[stat].min(), dataframe[stat].max(), dataframe[stat].max() - dataframe[stat].min()]
            for stat in self.get_stat_column_names_from_dataframe(dataframe)
        }

        weighting = [stat[2] + 1 for stat in meta.values()]  # max - min + 1
        weighting_min = min(weighting)

        # add weighting to each stat
        meta = {key: value + [weighting[i] / weighting_min] for i, (key, value) in enumerate(meta.items())}

        # adjust Suit_Jetpack_Ignition as half will be worst desirable now
        meta["Suit_Jetpack_Ignition"][2] /= 2

        # add threshold for Suit_Jetpack_Ignition
        meta["Suit_Jetpack_Ignition"].append(meta["Suit_Jetpack_Ignition"][1] - meta["Suit_Jetpack_Ignition"][2])

        return meta

    # recalaculate perfection for Suit_Jetpack_Ignition Hi/Lo desirability
    @staticmethod
    def calculate_ignition_perfection(meta: typing.Dict[str, typing.List[float]], dataframe: pandas.DataFrame) -> pandas.Series:
        result = []
        for i in range(len(dataframe)):
            perfection = []
            weighting_total = 0

            # calculate perfection for each seed
            for stat_name, stat_meta in meta.items():
                stat_value = dataframe.loc[i, stat_name]
                if numpy.isnan(stat_value):
                    continue

                weight = stat_meta[3]
                weighting_total += weight

                p = 1.0
                if stat_name != "Suit_Jetpack_Ignition" or stat_value > stat_meta[4]:
                    p -= (stat_meta[1] - stat_value) / stat_meta[2]
                else:
                    p -= (stat_meta[0] - stat_value) * -1 / stat_meta[2]
                perfection.append(p * weight)

            # add calculated perfection
            result.append((sum(perfection) / weighting_total) * (len(perfection) / 4))

        return pandas.Series(result)

    @staticmethod
    def remove_undesirable_weapon_boltcater_stats(dataframe: pandas.DataFrame) -> pandas.DataFrame:
        dataframe_desirable = dataframe[dataframe["Weapon_Projectile_Rate"].isna()]
        return dataframe_desirable[dataframe_desirable["Weapon_Projectile_ReloadTime"].isna()]

    @staticmethod
    def remove_undesirable_suit_movement_system_stats(dataframe: pandas.DataFrame) -> pandas.DataFrame:
        dataframe_desirable = dataframe[dataframe["Suit_Jetpack_Drain"].isna()]
        return dataframe_desirable[dataframe_desirable["Suit_Stamina_Strength"].isna()]

    # add the remaining best without duplicates until the target size is reached
    @staticmethod
    def fill_without_duplicates(dataframe_fill: pandas.DataFrame, dataframe_source: pandas.DataFrame, max_rows: int) -> pandas.DataFrame:
        dataframe_fill.drop_duplicates(subset="Seed", inplace=True)

        if  len(dataframe_fill) > max_rows:
            return dataframe_fill.sort_values(by=["Perfection", "Seed"], ascending=False)  # just sort here

        dataframe_remaining = dataframe_source[~dataframe_source["Seed"].isin(dataframe_fill["Seed"].values)]  # is not in
        head = dataframe_remaining.sort_values(by=["Perfection", "Seed"], ascending=False).head(int(max_rows - len(dataframe_fill)))  # take difference to max_rows

        return pandas.concat([dataframe_fill, head]).sort_values(by=["Perfection", "Seed"], ascending=False)  # sort by perfection for end result

    # endregion

    # region Sheets

    def generate_overview_sheet(self) -> None:
        # use first sheet for the overview
        sheet = self.workbook.active

        # set new title
        sheet.title = "Overview"

        # add description and prepare links to the items
        overview = [
            ["Every Item Procedural", "by zencq"],
            ["Version", VERSION],
            [],
            ["This file is a automatically generated compilation of all files available at GitHub in the repository zencq/Pi."],
            ["The repository contains a collection of files with values of stats for every procedural item in No Man's Sky. This includes mainly technology upgrades but also the products (treasures/artifacts in-game)."],
            ["Link to GitHub", URL],
            [],
            [],
            [],
            ["Technologies", "", "", "", "Products"],
        ]
        for row in overview:
            sheet.append(row)

        # make first line bold
        for column in sheet.iter_cols(min_col=1, max_col=2, min_row=1, max_row=1):
            for cell in column:
                cell.font = FONT_BOLD

        # autofit column width
        sheet.column_dimensions["A"].width = len(str(sheet["A1"].value))

        # set data format of version cell
        sheet["B2"].number_format = numbers.FORMAT_TEXT

        # set hyperlink attributes to url cell
        self.cell_set_hyperlink(sheet["B6"], URL)

        # merge reference cells as header
        self.row_max_merge_columns(sheet, 1, 3)  # technology
        self.row_max_merge_columns(sheet, 5, 8)  # product

        # set link for product
        self.cell_set_hyperlink(sheet.cell(row=sheet.max_row, column=5), f"{F_BASE}#'Product'!A1")

        # store data for to properly add links later
        self.link_settings = {
            "column_product": len(overview[-1]),
            "column_technology": 1,
            "row_first": len(overview) + 1,
        }

    def insert_in_overview_sheet(self, inventory_translation: str, item_id: str, target_coordinate = None) -> None:
        is_technology = inventory_translation != "Product"

        column = self.link_settings["column_technology" if is_technology else "column_product"]
        item_translation = _(item_id)
        row = self.link_settings["row_first"] + (TECHNOLOGY_INDEX.index(item_id) if is_technology else PRODUCT.index(item_id))
        sheet = self.workbook["Overview"]

        cell = sheet.cell(row=row, column=column, value=f"{inventory_translation}, {item_translation}")

        if target_coordinate:
            self.cell_set_hyperlink(cell, f"{F_BASE}#'{inventory_translation}'!{target_coordinate}")

    def insert_product(self, sheet: Worksheet, item_name: str, product: pandas.DataFrame) -> str:
        item_translation = _(item_name)

        # insert item_name
        cell = sheet.cell(row=sheet.max_row + 1, column=1, value=f"{item_translation} ({item_name})")
        # merge columns of row with item_name
        self.row_max_merge_columns(sheet, 1, 3)
        self.row_max_merge_columns(sheet, 4, 5)
        # make item_name bold
        cell.font = FONT_BOLD

        if len(product):
            for i, row in enumerate(dataframe_to_rows(product)):
                if i == 0:
                    index_name = row.index(self.language)
                    index_seed = row.index("Seed")
                    index_perfection = row.index("Perfection")
                    index_age = row.index("Age")
                    index_value = row.index("Value")
                    continue

                if i == 1:  # ignore as it is empty
                    continue

                # add data
                sheet.append([
                    row[index_name],
                    row[index_seed],
                    row[index_perfection],
                    row[index_age],
                    row[index_value],
                ])

                sheet.cell(row=sheet.max_row, column=3).number_format = "0%"  # perfection
        else:
            # add note that product is not available
            if item_name in OUTDATED:
                value = f"This product is not yet available in game version {VERSION} but was added in {OUTDATED[item_name]}."
            else:
                value = f"This product is not yet available in game version {VERSION}."

            self.cell_set_fill_color(sheet.cell(row=sheet.max_row + 1, column=1, value=value), FILL_RED)
            self.row_max_merge_columns(sheet, 1, 5)

        sheet.append([""])

    def insert_technology(self, sheet: Worksheet, item_id: str, data: typing.Dict[str, pandas.DataFrame]) -> str:
        stats_raw = [stat for stat in self.get_stat_column_names_from_dataframe(data.values())]

        # insert header
        sheet.append([f"{_(item_id)} ({item_id})", "Seed", "Perfection"] + [_(stat) for stat in stats_raw])
        # cache first row to later add it as reference in the overview sheet
        anchor_cell = sheet.cell(row=sheet.max_row, column=1)
        # make technology name bold
        anchor_cell.font = FONT_BOLD

        if not data:
            # add note that technology is not available
            if item_id in OUTDATED:
                value = f"This technology is not yet available in game version {VERSION} but was added in {OUTDATED[item_id]}."
            else:
                value = f"This technology is not yet available in game version {VERSION}."
            self.cell_set_fill_color(sheet.cell(row=sheet.max_row, column=4, value=value), FILL_RED)

            sheet.cell(row=sheet.max_row, column=4).alignment = ALIGNMENT_CENTER
            self.row_max_merge_columns(sheet, 4, 9)
        else:
            for (item_name, max_stats_per_seed, is_all_same), dataframe in data.items():
                quality_translation, colors = self.get_quality_information(item_name)

                outdated_item = item_id in OUTDATED and item_id or item_name in OUTDATED and item_name
                if outdated_item:
                    # add note that technology is outdated
                    sheet.append([quality_translation, "", max_stats_per_seed, f"This technology is outdated in game version {VERSION} as it was changed in {OUTDATED[outdated_item]}."])
                    self.row_max_set_fill_color(sheet, colors[1])
                    self.cell_set_fill_color(sheet.cell(row=sheet.max_row, column=4), FILL_RED)
                else:
                    sheet.append([quality_translation, "", max_stats_per_seed, "All seeds are the same." if is_all_same else ""])
                    self.row_max_set_fill_color(sheet, colors[1])

                sheet.cell(row=sheet.max_row, column=3).alignment = ALIGNMENT_LEFT
                sheet.cell(row=sheet.max_row, column=4).alignment = ALIGNMENT_CENTER
                self.row_max_merge_columns(sheet, 4, 9)

                for i, row in enumerate(dataframe_to_rows(dataframe)):
                    if i == 0:
                        index_name = row.index(self.language)
                        index_seed = row.index("Seed")
                        index_perfection = row.index("Perfection")
                        index_stats = {stat: row.index(stat) for stat in stats_raw}
                        continue

                    if i == 1:  # ignore as it is empty
                        continue

                    # add data
                    sheet.append(
                        [
                            row[index_name],
                            row[index_seed],
                            row[index_perfection],
                        ] + [row[index_stats[stat]] for stat in stats_raw]
                    )
                    # style added row
                    if i % 2 == 1:
                        self.row_max_set_fill_color(sheet, colors[0])

                    sheet.cell(row=sheet.max_row, column=3).number_format = "0.000%"

                    for column in sheet.iter_cols(min_col=4, max_col=4 + len(stats_raw) - 1, min_row=sheet.max_row, max_row=sheet.max_row):
                        for cell in column:
                            cell.number_format = "#,##0.00000"
                            # add darker color for best value per stat
                            if cell.value == dataframe[stats_raw[cell.col_idx - 4]].max():
                                self.cell_set_fill_color(cell, colors[1])

        self.row_max_set_border_bottom(sheet)
        sheet.append([""])

        return anchor_cell.coordinate

    # endregion

    # region Pandas

    def get_product_with_pandas(self, item_name) -> pandas.DataFrame:
        f_name = f"{F_PATH}\\Product\\{item_name}.parquet"
        if os.path.isfile(f_name):
            data = pandas.read_parquet(f_name)
            return data[data["Perfection"] == 1.0]

        return pandas.DataFrame()

    def get_technology_with_pandas(self, inventory_type, item_id, qualities: typing.List[str]) -> typing.Dict[str, pandas.DataFrame]:
        result = {}

        for i, quality in enumerate(qualities, 1):
            item_name = f"{item_id}{quality}"

            f_name = f"{F_PATH}\\{inventory_type}\\{item_name}.parquet"
            if not os.path.isfile(f_name):
                continue

            data = pandas.read_parquet(f_name)
            n = math.floor(3 / i)  # 3 for best, 1 for second

            columns = self.get_stat_column_names_from_dataframe(data)

            available_stats = len(columns)
            f_result, max_stats_per_seed, is_all_same = self.get_best_per_stat(data)

            if not is_all_same:
                dataframe_source = data

                # special desirability for certain items
                if item_id == "UP_JET":
                    # * DHarhan: for Movement, Recharge Rate + Sprint Recovery are more desirable than Fuel Effic + Sprint Dist
                    # * DHarhan: Movemement mods with lowest Initial Boost bonus are just as desirable as highest
                    # low bonus only for X as 4 always has high values
                    if quality == "X":
                        meta =  self.prepare_ignition_meta(data)
                        dataframe_source = dataframe_source.assign(Perfection=lambda row: self.calculate_ignition_perfection(meta, row))
                        dataframe_desirable = self.remove_undesirable_suit_movement_system_stats(dataframe_source)

                        # add 1/3 each for low and high bonus
                        ignition_l = dataframe_desirable[dataframe_source["Suit_Jetpack_Ignition"] < meta["Suit_Jetpack_Ignition"][4]].sort_values(by=["Perfection", "Seed"], ascending=False).head(n)
                        ignition_h = dataframe_desirable[dataframe_source["Suit_Jetpack_Ignition"] > meta["Suit_Jetpack_Ignition"][4]].sort_values(by=["Perfection", "Seed"], ascending=False).head(n)

                        f_result = pandas.concat([f_result, ignition_l, ignition_h])

                    dataframe_source = self.remove_undesirable_suit_movement_system_stats(data)  # do with original data
                    columns.remove("Suit_Jetpack_Drain")
                    columns.remove("Suit_Stamina_Strength")

                if item_id == "UP_BOLT":
                    # * DHarhan: Boltcaster mods with Dmg, Clip, ShotsPer, & BurstCool are desirable due to pairing well with ? mods with Dmg, Rate, Reload
                    dataframe_source = self.remove_undesirable_weapon_boltcater_stats(dataframe_source)
                    columns.remove("Weapon_Projectile_Rate")
                    columns.remove("Weapon_Projectile_ReloadTime")

                if item_id == "UP_GREN":
                    # * DHarhan: for PlasmaGren, the most desirable mods are just Damage+Velocity bonuses, followed by Dmg+Vel+Radius, never bounce
                    dataframe_source = dataframe_source[dataframe_source["Weapon_Grenade_Bounce"].isna()]
                    columns.remove("Weapon_Grenade_Bounce")
                    # * DHarhan: C Class P.Gren mods are better than S/A/B, because A forces an undesirable stat on it and S forces both unwanteds.
                    if quality == "1":
                        dataframe_source = dataframe_source[dataframe_source["Weapon_Grenade_Radius"].isna()]
                        columns.remove("Weapon_Grenade_Radius")

                # add the overall best per stat that is notna if max stats per seed is less than possible stats
                if available_stats > max_stats_per_seed:
                    for column in columns:
                        best = dataframe_source[dataframe_source[column].notna()].sort_values(by=["Perfection", "Seed"], ascending=False).head(1)
                        f_result = pandas.concat([f_result, best])

                    n *= HIGH_NUMBER_MULTIPLIER

                f_result = self.fill_without_duplicates(f_result, dataframe_source, n + available_stats)

            result[(item_name, max_stats_per_seed, is_all_same)] = f_result

        return result

    # endregion


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Generates a compilation of all files available in this repository.")

    parser.add_argument("-d", "--debug", action="store_true", help="Enable debug mode.")
    parser.add_argument("-l", "--language", choices=[RE_LANGUAGE.findall(l)[0] for l in LANGUAGES], default="en", help="Language to use for procedural names.")

    pi = parser.parse_args(namespace=Pi())

    pi.generate_overview_sheet()
    pi.add_technologies()
    pi.add_products()
    pi.write_xlsx()
